import os
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Alignment

def swap_elements(arr1, arr2, p_mut_cruza, p_mut_ind):
    if len(arr1) < 3 or len(arr2) < 3 or p_mut_cruza < 50:
        if p_mut_cruza < 50:
            print("\n[AUXILIAR WHEEL]\n\tThe cross mutation percentage is below 50%. No changes were made.")
        else:
            print("\n[AUXILIAR WHEEL]\n\tArrays must have at least 3 elements for swapping")
        return arr1, arr2

    percentage_per_element = 100 // (len(arr1))
    print("\n\t[AUXILIAR WHEEL]\n\tPERCENTAGE PER ELEMENT: [", percentage_per_element, "% ]")
    
    if(percentage_per_element <= p_mut_ind):
        print("\tRULE SET:", len(arr1), "elements could be crossed\n")
    else:
        print(f"\tPercentage calculated [{percentage_per_element} %] is below individual mutation percentage parameter [{p_mut_ind} %]. Changes could not be resolved due it\n")
        return arr1, arr2
    
    is_there_a_pair = False
    pairs_found = 0
    percentage_applied = percentage_per_element

    for i in range(len(arr1)):

        if percentage_applied <= p_mut_ind:
            # Check if the element at position i is equal in both arrays
            if arr1[i] == arr2[i]:
                if i == 0:
                    is_there_a_pair = True
                    pairs_found += 1
                    print(f"\tPair, posicion: {i}")
                else: 
                    if is_there_a_pair:
                        is_there_a_pair = True
                        pairs_found += 1
                        print(f"\tPair, posicion: {i}")
                    else:
                        is_there_a_pair = False
                pass
            else:
                if is_there_a_pair:
                    arr1[i], arr2[i] = arr2[i], arr1[i]
                else:
                    pass
                
            print("\tPercentage completed:", percentage_applied, " %")
            print(f"\t{arr1[i]} <-> {arr2[i]}, mix each other: {'Same value' if arr1[i] == arr2[i] else is_there_a_pair}\n")
            percentage_applied += percentage_per_element
        else:
            break
    if not is_there_a_pair:
        print("\tThere is no pair in the arrays. No changes were made.")
    print("\n\tPairs found:", pairs_found)
    return arr1, arr2

# Example usage
# arr1 = [1, 2, 3, 4, 5]
# arr2 = [5, 4, 3, 2, 1]
# arr1 = [0, 1, 3, 2, 2]
# arr2 = [2, 2, 3, 1, 0]
# arr1 = [1, 2, 3, 4, 5, 6, 8, 7, 9, 10, 11]
# arr2 = [1, 2, 3, 4, 5, 7, 6, 8, 9, 10, 11]
# arr1 = ['1-P1', '2-P2', '3-P3', '5-P4', '0-P5', '4-P6']
# arr2 = ['1-P1', '2-P2', '0-P3', '3-P4', '5-P5', '4-P6']

# print("Initially:")
# print("First pair:", arr1)
# print("Second pair:", arr2)

# arr1, arr2 = swap_elements(arr1, arr2, 50, 100)

# print("\nAt the end:")
# print("First pair:", arr1)
# print("Second pair:", arr2)

def define_operator(selected_option):
    if selected_option == 'a' or selected_option == 'c' or selected_option == 'e':
        print("[MUTATION] RULE SET: Ordening from lower to greater")
        return "<"
    else:
        print("[MUTATION] RULE SET: Ordening from greater to lower")
        return ">"

def mutation_for_array_parameter(array_parameter, operator, p_mut_gen, percentage_per_element, jumps):
    percentage_applied = 0
    changed_applies = 0
    
    print("\t[AUXILIAR WHEEL]")
    for i in range(jumps):
        percentage_applied += percentage_per_element
        
        if percentage_applied <= p_mut_gen:
            condition = "{} {} {}".format(array_parameter[i-changed_applies].split("-")[0], operator, array_parameter[len(array_parameter)-1].split('-')[0])
            print("\n\t[MUTATION] Condition:", condition)
            
            if eval(condition):
                print(f"\t[MUTATION] [X] No changes needed in {array_parameter[i-changed_applies]} <-> {array_parameter[len(array_parameter)-1]}")
            else:
                if i-changed_applies >= len(array_parameter)-1:
                    print(f"\t[MUTATION] [X] No changes made in {array_parameter[i-changed_applies]} <-> {array_parameter[len(array_parameter)-1]}")
                else:
                    print(f"\t[MUTATION] [✔] Changing {array_parameter[i-changed_applies]} <-> {array_parameter[len(array_parameter)-1]}")
                    array_parameter.append(array_parameter[i-changed_applies])
                    array_parameter.pop(i-changed_applies)
                    changed_applies += 1
                    print("\t[MUTATION] [N] New first_array got:\n\t\t", array_parameter)
                    
            print("\t[MUTATION] Percentage completed:", percentage_applied, " %")
        else:
            break
        
    print(f"\n\t[MUTATION] >> Mutation process completed by reaching the mutation percentage parameter [{p_mut_gen} %]. \n\t[MUTATION] >> Completed percentage:", percentage_applied, "% from", "100 % (default)" )
    return array_parameter

def calculate_fitness(array, operator, percentage_per_element):
    fitness_percentage = 100/(len(array)-1)
    fitness = 0
    
    for i in range(len(array)-1):
        condition = "{} {} {}".format(array[i].split("-")[0], operator, array[i+1].split('-')[0])
        print(f"\t* Condition", condition, f": is : {eval(condition)}")
        if eval(condition):
            fitness += fitness_percentage
        else:
            pass
    print("\t[AUXILIAR WHEEL]")
    print("\t[FITNESS] Fitness calculated:", fitness, "%\n")
    return fitness

def save_generation(name_file_report, arr1=[], arr2=[], current_epoch=0, fitness_arr1=0, fitness_arr2=0):
    # Ruta completa del archivo Excel
    print(f"\n[AUXILIAR WHEEL]\n[SAVE GENERATION] Saving the generation {current_epoch} in the file {name_file_report}")
    excel_file_path = os.path.join("reports", name_file_report)

    if os.path.exists(excel_file_path):
        # Cargar el archivo Excel existente
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active

        # Definir la fila de inicio para insertar los datos
        start_row = 5

        # Calcular el índice de la fila actual para arr1
        current_row_arr1 = start_row + current_epoch * 2

        # Calcular el índice de la fila actual para arr2
        current_row_arr2 = current_row_arr1 + 1

        # Insertar el encabezado si es la primera iteración
        if current_epoch == 0:
            headers = ["Num", "Result", "Fitness", "Epoch"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=start_row - 1, column=col, value=header)
                # Alinear el encabezado al centro
                ws.cell(row=start_row - 1, column=col).alignment = Alignment(horizontal='center', vertical='center')

        arr1_str = ', '.join(arr1)
        arr2_str = ', '.join(arr2)
        
        if arr1:
            # Insertar los datos de arr1 en las columnas A, B, C y D respectivamente
            ws.cell(row=current_row_arr1, column=1, value=current_row_arr1-4).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row_arr1, column=2, value=arr1_str).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row_arr1, column=3, value=fitness_arr1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row_arr1, column=4, value=current_epoch).alignment = Alignment(horizontal='center', vertical='center')

        if arr2:
            # Insertar los datos de arr2 en las columnas A, B y C respectivamente
            ws.cell(row=current_row_arr2, column=1, value=current_row_arr2-4).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row_arr2, column=2, value=arr2_str).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row_arr2, column=3, value=fitness_arr2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row_arr2, column=4, value=current_epoch).alignment = Alignment(horizontal='center', vertical='center')
            
        # Ajustar el ancho de las columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener el nombre de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Guardar el archivo Excel
        wb.save(excel_file_path)
    else:
        messagebox.showerror("Error", f"El archivo {name_file_report} no existe. \nPor favor, verifica la existencia del archivo.")

def save_best_generation(name_file_report, best_array, fitness_reached, epoch):
    # Ruta completa del archivo Excel
    print(f"\n[AUXILIAR WHEEL]\n[SAVE GENERATION] Saving the best generation {epoch} in the file {name_file_report}")
    excel_file_path = os.path.join("reports", name_file_report)

    if os.path.exists(excel_file_path):
        # Cargar el archivo Excel existente
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active

        # Definir la fila de inicio para insertar los datos
        start_row = 1
        start_column = 7

        # Cabeceras
        headers = ["Epoch", "Best Result", "Fitness reached"]
        for col, header in enumerate(headers, start=start_column):
            ws.cell(row=start_row, column=col, value=header)
            # Alinear el encabezado al centro
            ws.cell(row=start_row, column=col).alignment = Alignment(horizontal='center', vertical='center')

        arr1_str = ', '.join(best_array)
        
        # Insertar los datos
        ws.cell(row=2, column=start_column, value=epoch-4).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=2, column=start_column + 1, value=arr1_str).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=2, column=start_column + 2, value=fitness_reached).alignment = Alignment(horizontal='center', vertical='center')

        # Ajustar el ancho de las columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener el nombre de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Guardar el archivo Excel
        wb.save(excel_file_path)
    else:
        messagebox.showerror("Error", f"El archivo {name_file_report} no existe. \nPor favor, verifica la existencia del archivo.")