import openpyxl
from tkinter import filedialog
import pandas as pd
import os
from datetime import datetime
from tkinter import messagebox

file_name_made = ""
file_route = ""

def open_excel_file():
    global file_route
    # Obtener la ruta del directorio del script actual
    project_route = os.path.dirname(__file__)
    
    # Unir la ruta del proyecto con el nombre de la carpeta "excel"
    default_route = os.path.join(project_route, "excel_files")

    # Verificar si la ruta predeterminada es válida
    if os.path.exists(default_route):
        file_route = filedialog.askopenfilename(initialdir=default_route, filetypes=[("Archivos Excel", "*.xlsx;*.xls")])
    else:
        print("[EXCEL WHILE] The default route does not exist.")
        file_route = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx;*.xls")])
    # Verificar si se seleccionó un archivo
    if file_route:
        # Leer el archivo Excel seleccionado usando pandas
        try:
            df = pd.read_excel(file_route)
            print("[EXCEL WHILE] Excel file read successfully.")
            print(df)
            process_excel_file(df)
            return {"success": "Archivos creados exitosamente en la carpeta 'temp_files'."}
        except Exception as e:
            print(f"[EXCEL WHILE] Error reading the Excel file: {e}")
            return {"error": e}
    else:
        print("[EXCEL WHILE] No file selected.")
        
def process_excel_file(df_file):
    if not os.path.exists('temp_files'):
        os.makedirs('temp_files')
        
    # Crear el primer archivo "duration"
    df_duration = df_file[['Nombre de curso', 'Duración (en minutos)']].rename(columns={'Nombre de curso': 'id', 'Duración (en minutos)': 'level'})
    df_duration.to_excel('temp_files/duration.xlsx', index=False)

    # Crear el segundo archivo "difficulty"
    df_difficulty = df_file[['Nombre de curso', 'Dificultad']].rename(columns={'Nombre de curso': 'id', 'Dificultad': 'level'})
    df_difficulty['conversion_num'] = df_difficulty['level'].apply(convert_difficulty)
    df_difficulty.to_excel('temp_files/difficulty.xlsx', index=False)

    # Crear el tercer archivo "requirements"
    df_requirements = df_file[['Nombre de curso', 'Requisitos']].rename(columns={'Nombre de curso': 'id', 'Requisitos': 'size'})
    df_requirements['conversion_num'] = df_requirements['size'].apply(lambda x: 0 if pd.isna(x) else len(str(x).split(',')))
    df_requirements.to_excel('temp_files/requirements.xlsx', index=False)

    center_excel_content('temp_files/duration.xlsx')
    center_excel_content('temp_files/difficulty.xlsx')
    center_excel_content('temp_files/requirements.xlsx')
    print("\n[EXCEL WHILE] Files created successfully in the 'temp_files' folder.")

def convert_difficulty(difficulty):
    difficulty = str(difficulty).lower()
    if difficulty in ['facil', 'fácil', '0', 'easy']:
        return 0
    elif difficulty in ['media', 'medio', '1', 'medium']:
        return 1
    elif difficulty in ['difícil', 'dificil', '2', 'hard']:
        return 2
    elif difficulty in ['muy difícil', 'muy dificil', '3', 'very hard']:
        return 3
    elif difficulty in ['insano', 'maestro', '4', 'insane', 'master']:
        return 4
    else:
        return 0

def save_default_data_to_excel(data):
    global file_name_made
    # Get the current date and time
    now = datetime.now()
    date = now.strftime("%d-%m-%Y")
    time = now.strftime("%H%M%S")

    # Create the file name
    file_name = f"report_{date}_{time}.xlsx"
    file_name_made = file_name

    # Create the folder if it doesn't exist
    if not os.path.exists("reports"):
        os.makedirs("reports")

    # Create the Excel workbook
    wb = openpyxl.Workbook()

    # Select the active sheet
    ws = wb.active

    # Write the data to the Excel file
    ws.append(["Order by", "Epochs", "P_mut_cruza", "P_mut_ind", "P_mut_gen"])
    ws.append([data["selected_option"], data["iterator"], data["p_mut_cruza"], data["p_mut_ind"], data["p_mut_gen"]])

    # Save the workbook
    wb.save(os.path.join("reports", file_name))
    center_excel_content(os.path.join("reports", file_name))
    return file_name
    

def center_excel_content(file_path):
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Centrar el contenido de todas las celdas
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        wb.save(file_path)
    else:
        print(f"El archivo {file_path} no existe.")
        
def open_excel_made():
    excel_file_path = os.path.join("reports", file_name_made)
    
    if os.path.exists(excel_file_path):
        os.startfile(excel_file_path)
        return
    else:
        messagebox.showerror("EXCEL WHEEL", f"Ocurrió un error inesperado. \nNo podemos encontrar el archivo Excel guardado.\n{file_name_made}")
        return
    
def crear_excel_cursos_ordenados():
    global file_name_made, file_route
    print("[EXCEL WHILE] Creando archivo Excel con cursos ordenados...")
    print("[EXCEL WHILE] Extracting report file: ", file_name_made)
    print("[EXCEL WHILE] Extracting course file: ", file_route)
    
    # Crear la carpeta "final courses" si no existe
    final_courses_folder = "final courses"
    if not os.path.exists(final_courses_folder):
        os.makedirs(final_courses_folder)
    
    # Leer los datos del archivo "report"
    report_file_path = os.path.join("reports", file_name_made)
    wb_report = openpyxl.load_workbook(report_file_path)
    ws_report = wb_report.active
    
    # Obtener los valores de la columna H (segunda columna)
    column_H_values = [cell.value for cell in ws_report['H'][1:]]
    
    # Filtrar los valores None de la lista
    column_H_values_filtered = [value for value in column_H_values if value is not None]
    
    # Convertir la lista de valores en una cadena de texto separada por comas
    cursos_ordenados_str = ', '.join(map(str, column_H_values_filtered))
    
    print("[EXCEL WHILE] Cursos ordenados: ", cursos_ordenados_str)
    
    # Leer los datos del archivo de cursos
    wb_course = openpyxl.load_workbook(file_route)
    ws_course = wb_course.active

    # Crear una lista de nombres de cursos ordenados
    nombres_cursos_ordenados = []
    # Crear un diccionario para almacenar las posiciones y nombres de los cursos
    posiciones_nombres = {}
    
    # Iterar sobre las filas de la columna A desde la fila 2 en adelante
    for row_index, row in enumerate(ws_course.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
        nombre_curso = row[0]
        # Verificar si nombre_curso no es None antes de agregarlo al diccionario
        if nombre_curso is not None:
            posiciones_nombres[row_index - 1] = nombre_curso
    
    # Organizar los nombres de los cursos según las posiciones dadas en cursos_ordenados_str
    nuevos_datos_cursos = []
    for nombre in cursos_ordenados_str.split(", "):
        posicion_str = nombre.split("-")[1]
        if posicion_str.startswith("P"):
            posicion = int(posicion_str[1:])
            nuevo_nombre = posiciones_nombres.get(posicion)
            if nuevo_nombre:
                nuevos_datos_cursos.append((f"{nombre}", nuevo_nombre))
                nombres_cursos_ordenados.append((f"P{posicion}", nuevo_nombre))
    
    # Guardar los datos ordenados en un nuevo archivo Excel
    now = datetime.now()
    date_string = now.strftime("%d_%m_%Y_%H_%M_%S")
    file_name = f"final_course_{date_string}.xlsx"
    final_course_path = os.path.join(final_courses_folder, file_name)
    
    wb_final_course = openpyxl.Workbook()
    ws_final_course = wb_final_course.active
    
    # Escribir los encabezados
    ws_final_course["A1"] = "Posición"
    ws_final_course["B1"] = "Nombre de curso"
    
    # Escribir los datos de los cursos ordenados
    for i, (posicion, nombre) in enumerate(nombres_cursos_ordenados, start=2):
        ws_final_course[f"A{i}"] = posicion
        ws_final_course[f"B{i}"] = nombre
    
    # Guardar el archivo Excel final
    wb_final_course.save(final_course_path)
    
    print(f"[EXCEL WHILE] Archivo Excel con cursos ordenados creado: {final_course_path}")